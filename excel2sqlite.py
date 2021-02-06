''' #!/usr/bin/python3.7 '''
#!~/Projects/venvs/kivy_py37/bin/python3.7

import os
import sys
import logging
import pysqlite3 as sqlite3
import diagnostics as diag
from openpyxl import load_workbook


__title__ = sys.argv[0].split('/')[-1]
__author__ = "Mccoyflatline"
__copyright__ = "Copyright 2020 lololol"
__credits__ = ["Mccoyflatline"]
__license__ = "Not Sure Yet"
__version__ = "0.9"
__maintainer__ = "Mccoyflatline"
__email__ = "pitft@protonmail.com"
__status__ = "Pre-release"


# if no log folder, create one.
if not os.path.exists('logs/'):
    os.mkdir('logs')

# initialize logger
logging.basicConfig(filename='logs/db_log.log', level=logging.NOTSET,
                    format=':%(asctime)s-%(levelname)s-%(lineno)s-%(message)s')


######
## OPENPYXL
## GRAB DATA FROM EXCEL FILE

# Grab worksheet data
def grab_worksheet(xcel_filename, wkst_name):
    ''' Loads excel workbook then returns Worksheet class object. '''
    wb = load_workbook(filename=xcel_filename, keep_links=False)
    return wb[wkst_name]


# access individual cells by --> wkst.cell(row,col)
# first_row --> wk.cell(1,1) == 'A1', wk.cell(2,1) == 'A2'
# first_col --> wk.cell(1,1) == 'A1', wk.cell(1,2) == 'B1'
def grab_fields(wkst, last_col, first_col=1, row=1, skip=[None, ' ']):
    ''' Returns column names [fields] from excel worksheet by
        looping through worksheet cells then grabbing the value. '''
    fields = []
    # add 1 to include the last field
    last_col += 1
    for col in range(first_col, last_col):
        if wkst.cell(row,col).value in skip or col in skip:
            continue
        fields.append(wkst.cell(row,col).value)
    return fields


# grab datatypes / data formats
def grab_types(wkst, last_col, first_col=1, row=1, skip=[None, ' ']):
    ''' Returns datatypes [types] (number formatting in excel) of
        cell values. '''
    types = []
    for col in range(first_col, last_col+1):
        if wkst.cell(row,col).value in skip or col in skip:
            continue
        types.append(wkst.cell(row,col).number_format)
    return types


# combine fields into a schema string for an sqlite statement
def create_schema(fields, types, pri_key=False):
    ''' Loops through and concatenates [fields] and [types] then
        returns 'schema' as a string. '''
    schema=[]
    for num in range(len(fields)):
        # Debating just putting REAL since that's what sqlite does.
        if '.' in types[num]:
            types[num] = 'DOUBLE'
        elif '0' in types[num]:
            types[num] = 'INTEGER'
        elif types[num] == 'General':
            types[num] = 'TEXT'
        else:
            print('Unidentified number_format: ' + str(fields[num]) + " " + str(types[num]))
        # Adding option for primary, keeping commented out for now
        # assuming the column number is passed through pri_key
#        if num == pri_key:
#               types[pri_key] +=' PRIMARY KEY'

        schema.append(fields[num] + ' ' + types[num])

    return ", ".join(schema)


def grab_records(wkst, last_col, last_row, first_col=1, first_row=1, skip=[]):
    ''' Returns dict() records, which is {"number": "row_of_cell_values"}.
        Use if memory usage is not significant. '''
    records = {}
    # Have to add one because cols and rows can't start at 0,
    # and has to include the actual last row and column
    last_col+=1
    last_row+=1
    #starting at row 1 to the nth row.
    for row_num in range(first_row, last_row):

        # if row is skippable, skip
        # [Blank product number, ' ', or ItenNum title]
        if wkst.cell(row_num,2).value in [None, ' ', "ItemNum"]:
            continue
        else:
            # else, prep the next row into the dictionary
            records.update({str(row_num):[]})

    # starting at 'A' (first col) to the nth col.
        for col_num in range(first_col, last_col):
            # if cell is skippable, skip.
            if col_num in skip:
                continue
            #if number formatted cell is empty, fill with 0
            if wkst.cell(row_num,col_num).value == None:
                wkst.cell(row_num,col_num).value = 0
            # append cell's value to list in the dictionary
            records[str(row_num)].append(stringify(wkst.cell(row_num,col_num).value))
    return records


##  Feb, 6, 2021. Thoughts:
##  remove skip for rows, keep for columns.
def grab_records_gen(wkst, last_col, last_row, first_col=1, first_row=1, skip=[]):
    ''' Yields dict() records, which is {"number": "row_of_cell_values"}.
        Use if memory usage is significant.'''
    records = {}
    # Have to add one because cols and rows can't start at 0,
    # and has to include the actual last row and column
    last_col+=1
    last_row+=1
    #starting at row 1 to the nth row.
    for row_num in range(first_row, last_row):

        # if row is skippable, skip
        # [Blank product number, ' ', or ItenNum title]
        if wkst.cell(row_num,2).value in [None, ' ', "ItemNum"]:
            continue
        else:
            # else, prep the next row into the dictionary
            records.update({str(row_num):[]})

    # starting at 'A' (first col) to the nth col.
        for col_num in range(first_col, last_col):
            # if column is skippable, skip.
            if col_num in skip:
                continue

            #if number formatted cell is empty, fill with 0
            if wkst.cell(row_num,col_num).value == None:
                wkst.cell(row_num,col_num).value = 0

            # append cell's value to list in the dictionary
            records[str(row_num)].append(stringify(wkst.cell(row_num,col_num).value))

        yield ', '.join(records[str(row_num)])
        records.clear()



# 1 -> 'A', 27 -> 'AA', 455 -> 'QM'
def grab_col_letter(col_number):
    ''' Returns column letter name in relation to the column's
        number (col_number).
    '''
    alphabet = {0:'Z',1:'A',2:'B',3:'C',4:'D',5:'E',6:'F',7:'G',8:'H',
                9:'I',10:'J',11:'K',12:'L',13:'M',14:'N',15:'O',
                16:'P',17:'Q',18:'R',19:'S',20:'T',21:'U',22:'V',
                23:'W',24:'X',25:'Y',26:'Z'}

    result=''
    # Assuming col_number is never higher than 1024
    ''' As of 01:38AM on Feb 6, 2021, I finally fixed the math on this
        freaking function.

        Sadly, I don't feel as happy as I should. I don't feel happy at all.
        The amount of head banging the wall won't make up for the lack of
        sleep.

        If you're stuck on a problem, set a timer for 10-15 minutes.
        If you can't solve the problem in that time, take a break or,
        in my case, GO TO SLEEP.
    '''
    while col_number > 0:
        result += alphabet[col_number % 26]
    # imagine calling the same variable 3 times in the same short equation.
        col_number -= col_number - (col_number // 26)
    ''' The math above returns the correct letters right to left first,
        so just reversing it makes it a lot simpler to deal with.
    '''
    return result[::-1]


# 'A' -> 1, 'QM' -> 455, 'BRC'-> 1823
def grab_col_index(column):
    ''' Return the number index of the column. '''
    betalpha = {' ':0,'A':1,'B':2,'C':3,'D':4,'E':5,'F':6,'G':7,'H':8,
                'I':9,'J':10,'K':11,'L':12,'M':13,'N':14,'O':15,
                'P':16,'Q':17,'R':18,'S':19,'T':20,'U':21,'V':22,
                'X':23,'Y':24,'X':25,'Z':26}
    # Reversing the string makes this a whole heck of a lot easier
    column = column[::-1]
    index = 0
    for num in range(len(column)):
        index += (betalpha[column[num].upper()] % 27 * (26**num))
    return index


def find_length_row(wkst, row=1):
    ''' Find the number of cells between:
        Row's 1st cell that has a value and row's last cell that has a value.
    '''
    for col in range(1,(len(wkst[row])+1)):
        if wkst.cell(row, col).value != None:
            # have to add 1, otherwise it'll always be off by 1
            return ( len(wkst[row]) + 1) - col

#col can be a letter or number. 'A' --> 1, 'AB' --> 28
def find_length_col(wkst, col='A'):
    ''' Find the number of cells between:
        column's 1st cell that has a value and column's last cell that has a value.
    '''
    #if col is int, grab column letter.
    if isinstance(col, int):
        pass
#        col = grab_col_letter(col)
    # assuming col is str
    col_index = grab_col_index(col)

    for row in range(1,(len(wkst[col]) + 1)):
        if wkst.cell(row, col_index).value != None:
            # have to add 1, otherwise it'll always be off by 1
            return ( len(wkst[col]) + 1) - row


def table_exists(cursor, table):
    ''' Check to see if table exists. '''
    cursor.execute(f'''SELECT count(name) FROM sqlite_master
                    WHERE type='table' AND name='{table}' ''')

    # If the count is 1, then table exists
    if cursor.fetchone()[0] == 1:
        logging.debug(f"'{table}' exists.")
        return True
    else:
        logging.debug(f"'{table}' doesn't exist.")
        return False


# Create a table.
def create_table(cursor, table, schema):
    ''' Create sqlite table if it doesn't exist.'''
    # Check to see if table exists
    if table_exists(cursor, table):
        logging.debug(f"Table '{table}' already exists.")
        return  # breaking out of function
    else:
        # Schema - The field names and types of data the table is storing.
        cursor.execute(f'''CREATE TABLE {table} ({schema})''')
        logging.info(f"Table '{table} ({schema})' created.")


# assuming filename has no '.' other than extension if at all
# if filename doesn't have ext, add ext.
def check_extension(filename, ext='.db'):
    ''' Check if filename has the correct extension.
        Return True, if yes. Return False, if not.
    '''
    if filename[:-(len(ext))] == ext:
        return True
    else:
        return False

# This function is under the assumption that the only extensions we're
# dealing with are '.db' and 'xlsx' or similar.
# It's a good idea pass the '.' along with the extension name.
# Also, a good idea to not have filename using '.' as spaces.
def change_extension(filename, ext='.db'):
    ''' Change filename's extension to new extension '''
    # if extension is already correct, leave alone.
    if filename[-(len(ext)):] == ext:
        pass
    # if filename doesn't have a '.', add '.' + ext
    elif '.' not in filename:
        if '.' not in ext:
            filename += '.' + ext
        else:
            filename += ext
    # if extension is not correct, correct it.
    elif filename[:-len(ext)] != ext:
        if '.' in ext:
            filename = filename[:filename.find('.')] + ext
        else:
            filename += filename[:filename.find('.')] + '.' +ext
    return filename


# connects to database. returns connection object
# if database doesn't exist, sqlite3 creates it.
# So this will probably be the usual way to make
# new databases.
def connect(sqlite3_mod, database):
    ''' Create/Connect to database then grab the connection '''
    database = change_extension(database)
    return sqlite3_mod.connect(database)


# Have to use the cursor object, so that means
# you must already have an sqlite connection to use this.
# creates database, doesn't return anything.
def create_db(cursor, database):
    ''' Create a standalone database '''
    database = change_extension(database)
    cursor.execute(f'{database}')

    # note: I guss I could do this by using subprocess module to
    #       to create the process of using sqlite3 cli tool to
    #       create the database then kill process (if it doesn't
    #       do it itself), but I don't feel like doing that.


def connect_to_db(database):
    ''' Create/Connect to database then grab the connection and cursor '''
    database = change_extension(database)
    connection = connect(sqlite3, database)
    cursor = grab_cursor(connection)
    return connection, cursor


def grab_cursor(connection):
    ''' Return cursor object '''
    return connection.cursor()


def commit(connection):
    ''' Commit connection's transactions '''
    connection.commit()


# close connection
def close(connection):
    ''' Close connection '''
    connection.close()


# inserting data into some (or all, if you want) columns
def insert_cols(cursor, table, col_names, row_data, *args, **kwargs):
    ''' Insert data into table for specific columns. '''
    cursor.execute(f'''INSERT INTO {table} ({",".join(col_names)}) VALUES ({','.join(row_data)})''')


# inserting data to every column
def insert_all(cursor, table, row_data):
    ''' Insert data into table for all columns.'''
    cursor.execute(f'''INSERT INTO {table} VALUES ({row_data})''')


# Prepping select queries for sending sqlite database into 2010` excel files
# select from certain columns
def select(cursor, table, cols):
    ''' Select query of specific columns from table. '''
    cursor.execute(f'''SELECT {','.join(cols)} from {table}''')

# select all columns
def select_all(cursor, table):
    ''' Select query of all column from table. '''
    cursor.execute(f'''SELECT * from {table}''')


# sqlite really wants single quotes around text data.
# ex) s1 = 'string', s2 = '\'string\''
# cursor.execute(f"insert into table_a values ({s1})"
#   -> "insert into table_a values (string) x
#   -> Error: No column named string
# cursor.execute(f"insert into table_a values ({s2})"
#   -> "insert into table_a values ('string') ✔
def stringify(string):
    ''
    return (f'\'{str(string)}\'')

if __name__ == '__main__':
    # Basic program information
    diag.__info__(__title__, __version__, __status__, __author__, __email__)

    print('BRC',str(grab_col_index('BRC')))
    print(str(455),str(grab_col_letter(455)))
    print(str(79),str(grab_col_letter(79)))
